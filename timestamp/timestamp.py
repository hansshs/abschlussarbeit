import pandas as pd
import os
import openpyxl
import paths as ph

csv_paths       = ph.iter_path0311
output_folder   = ph.output_folder
elapsed_time    = []



for csv_path in csv_paths:
    df = pd.read_csv(csv_path)

    start_time = df['ElapsedSeconds'].iloc[0]
    end_time   = df['ElapsedSeconds'].iloc[-1]

    runtime = end_time - start_time
    elapsed_time.append({'Elapsed_Time [s]': runtime, 'Elapsed_Time [min]':f'{runtime/60: .3f}'})

df_result = pd.DataFrame(elapsed_time)
output_file = os.path.join(output_folder, ph.title[3])

wb = openpyxl.Workbook()
ws = wb.active

# Adicionar os títulos das colunas à planilha
column_titles = list(df_result.columns)
for col_idx, title in enumerate(column_titles, start=1):
    ws.cell(row=1, column=col_idx).value = title

# Escrever dados na planilha do Excel
for row_idx, row_data in enumerate(df_result.values, start=2):  # Comece da linha 2 para deixar espaço para os títulos
    for col_idx, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx).value = cell_data


# Save the Excel file
wb.save(output_file)
#df_result.to_csv(output_file, index=False)